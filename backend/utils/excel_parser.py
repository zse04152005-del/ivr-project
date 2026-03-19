"""Excel / CSV 导入解析工具"""
import re
import io
from typing import List, Tuple
from openpyxl import load_workbook


def parse_excel(file_content: bytes, filename: str) -> Tuple[List[dict], List[str]]:
    """
    解析上传的 Excel/CSV 文件，返回 (成功记录列表, 错误信息列表)

    Excel 格式要求：
    第1列：姓名
    第2列：电话
    第3列：年龄（可选）
    第4列：社区（可选）
    第一行为表头，从第二行开始读数据
    """
    records = []
    errors = []

    try:
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            name = str(row[0]).strip() if row[0] else ""
            phone = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            age = row[2] if len(row) > 2 else None
            community = str(row[3]).strip() if len(row) > 3 and row[3] else ""

            # 校验姓名
            if not name:
                errors.append(f"第{row_idx}行：姓名为空，已跳过")
                continue

            # 校验电话格式
            phone = re.sub(r"[^\d]", "", phone)  # 去除非数字字符
            if not phone or len(phone) < 7:
                errors.append(f"第{row_idx}行（{name}）：电话号码格式不正确「{phone}」")
                continue

            # 校验年龄
            if age is not None:
                try:
                    age = int(age)
                    if age < 0 or age > 150:
                        errors.append(f"第{row_idx}行（{name}）：年龄超出范围「{age}」，已设为空")
                        age = None
                except (ValueError, TypeError):
                    age = None

            records.append({
                "name": name,
                "phone": phone,
                "age": age,
                "community": community,
            })

        wb.close()
    except Exception as e:
        errors.append(f"文件解析失败：{str(e)}")

    return records, errors
