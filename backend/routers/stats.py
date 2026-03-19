"""统计报表路由"""
import io
import logging
from datetime import datetime
from urllib.parse import quote

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from models import Patient, CallTask, Appointment, CallStatus
from schemas import BatchStats

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/stats", tags=["统计报表"])

STATUS_LABELS = {
    "pending":     "待拨打",
    "calling":     "拨打中",
    "accepted":    "同意体检",
    "rejected":    "拒绝体检",
    "no_answer":   "未接/无响应",
    "to_schedule": "同意待约",
    "transferred": "已转人工",
    "failed":      "拨打失败",
}

HEADER_FILL = PatternFill("solid", fgColor="366092")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@router.get("/overview", summary="总览统计")
def overview_stats(db: Session = Depends(get_db)):
    """返回系统总览数据"""
    total_patients = db.query(Patient).count()
    total_calls = db.query(CallTask).filter(CallTask.call_count > 0).count()

    statuses = {}
    for status in CallStatus:
        statuses[status.value] = (
            db.query(CallTask).filter(CallTask.status == status.value).count()
        )

    connected = (
        statuses.get("accepted", 0)
        + statuses.get("rejected", 0)
        + statuses.get("to_schedule", 0)
        + statuses.get("transferred", 0)
    )
    accepted = (
        statuses.get("accepted", 0)
        + statuses.get("to_schedule", 0)
        + statuses.get("transferred", 0)
    )

    return {
        "total_patients": total_patients,
        "total_calls": total_calls,
        "connected": connected,
        "connect_rate": round(connected / total_calls * 100, 1) if total_calls else 0,
        "accepted": accepted,
        "accept_rate": round(accepted / connected * 100, 1) if connected else 0,
        "rejected": statuses.get("rejected", 0),
        "no_answer": statuses.get("no_answer", 0),
        "to_schedule": statuses.get("to_schedule", 0),
        "transferred": statuses.get("transferred", 0),
        "total_appointments": db.query(Appointment).count(),
    }


@router.get("/batch/{batch_id}", response_model=BatchStats, summary="单批次统计")
def batch_stats(batch_id: str, db: Session = Depends(get_db)):
    """返回某批次的详细统计"""
    tasks = db.query(CallTask).filter(CallTask.batch_id == batch_id).all()
    stats = {
        "total": len(tasks),
        "pending": 0, "calling": 0, "accepted": 0, "rejected": 0,
        "no_answer": 0, "to_schedule": 0, "transferred": 0, "failed": 0,
    }
    for t in tasks:
        if t.status in stats:
            stats[t.status] += 1
    return BatchStats(batch_id=batch_id, **stats)


@router.get("/export/{batch_id}", summary="导出Excel报表")
def export_batch_excel(batch_id: str, db: Session = Depends(get_db)):
    """导出某批次的拨打结果为Excel文件"""
    tasks = (
        db.query(CallTask)
        .filter(CallTask.batch_id == batch_id)
        .order_by(CallTask.id)
        .all()
    )

    wb = openpyxl.Workbook()

    # Sheet 1: 汇总
    ws1 = wb.active
    ws1.title = "批次汇总"
    _write_summary_sheet(ws1, batch_id, tasks)

    # Sheet 2: 明细
    ws2 = wb.create_sheet("拨打明细")
    _write_detail_sheet(ws2, tasks)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"IVR报表_{batch_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ===== 内部辅助函数 =====

def _set_header(ws, row: int, cols: list):
    for col_idx, name in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_summary_sheet(ws, batch_id: str, tasks: list):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    ws["A1"] = "批次ID"
    ws["B1"] = batch_id
    ws["A2"] = "导出时间"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "总人数"
    ws["B3"] = len(tasks)

    counts = {s: 0 for s in STATUS_LABELS}
    for t in tasks:
        if t.status in counts:
            counts[t.status] += 1

    row = 5
    _set_header(ws, row, ["状态", "数量", "占比"])
    ws.column_dimensions["C"].width = 12
    row += 1
    for status, label in STATUS_LABELS.items():
        count = counts[status]
        pct = f"{round(count / len(tasks) * 100, 1)}%" if tasks else "0%"
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=pct)
        row += 1


def _write_detail_sheet(ws, tasks: list):
    cols = ["序号", "姓名", "电话", "年龄", "社区", "状态", "拨打次数",
            "最近拨打时间", "按键记录", "是否转人工", "备注"]
    _set_header(ws, 1, cols)

    col_widths = [6, 12, 14, 6, 15, 12, 8, 20, 10, 10, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    for row_idx, task in enumerate(tasks, start=2):
        p = task.patient
        ws.cell(row_idx, 1, row_idx - 1)
        ws.cell(row_idx, 2, p.name if p else "")
        ws.cell(row_idx, 3, p.phone if p else "")
        ws.cell(row_idx, 4, p.age if p else "")
        ws.cell(row_idx, 5, p.community if p else "")
        ws.cell(row_idx, 6, STATUS_LABELS.get(task.status, task.status))
        ws.cell(row_idx, 7, task.call_count)
        ws.cell(row_idx, 8,
                task.called_at.strftime("%Y-%m-%d %H:%M:%S") if task.called_at else "")
        ws.cell(row_idx, 9, task.key_pressed or "")
        ws.cell(row_idx, 10, "是" if task.transferred else "否")
        ws.cell(row_idx, 11, task.notes or "")
